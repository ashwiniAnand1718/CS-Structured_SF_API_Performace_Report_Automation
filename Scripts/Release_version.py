import openpyxl
import os
# Please provide directory path of PT Reports
dir_PT_reports = "C:\\Users\\sa\\PycharmProjects\\PT_report_formatting\\Formatted_report\\"


pt_reports = ["MA-SF API_performance_report_credit_model_win64", "MA-SF API_performance_report_credit_model_lin64", "MA-SF API_performance_report_static_win64", "MA-SF API_performance_report_static_lin64"]
version = input(f"Please Provide Release Version : ")
for report in pt_reports:
    report_path = dir_PT_reports+report
    work_book = openpyxl.load_workbook(report_path + ".xlsx")
    work_sheet = work_book.worksheets[0]
    if "credit" in report:
        work_sheet.cell(row=6, column=3).value = version
    elif "static" in report:
        work_sheet.cell(row=5, column=3).value = version
    work_book.save(report_path + "_" + version + ".xlsx")
    os.remove(report_path+".xlsx")
    print(f"{report}_{version} done")


