# Please provide directory path of downloaded result
dwndload_res_path = "C:\\Users\\sa\\PycharmProjects\\PT_report_formatting\\Download_result\\"

# Please provide directory path of template
temp_path = "C:\\Users\\sa\\PycharmProjects\\PT_report_formatting\\Template\\"

# Please provide directory path to store final result
report_path = "C:\\Users\\sa\\PycharmProjects\\PT_report_formatting\\Formatted_report\\"


#===================================================================================================
import openpyxl
import os
import shutil
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment


# Script for Copying Template
source_folder = temp_path
destination_folder = report_path

# fetch all files
for file_name in os.listdir(source_folder):
    # construct full file path
    source = source_folder + file_name
    destination = destination_folder + file_name
    # copy only files
    if os.path.isfile(source):
        shutil.copy(source, destination)

## Script for Credit Model

OS = ["Windows", "Linux"]
copy_file_names = "analysis_results_all_summary_credit_model.xlsx"
paste_file_names = ["MA-SF API_performance_report_credit_model_win64.xlsx", "MA-SF API_performance_report_credit_model_lin64.xlsx"]
k = 0
for oss in OS:
    # Copy
    raw_file = dwndload_res_path+oss+"\\"+copy_file_names
    raw_workbook = openpyxl.load_workbook(raw_file)
    Summary = raw_workbook["Summary"]

    # Paste
    formatted_file = report_path+paste_file_names[k]
    formatted_workbook = openpyxl.load_workbook(formatted_file)
    Test_Summary = formatted_workbook["Test Summary"]


    copy_rows = [4,10,24,30,54,60,84,90,104,110]
    paste_rows = [11,17,22,28,33,39,44,50,55,61]
    a = 0
    b = 1
    z = 5
    while z >= 1:
        rangeselectted = []
        for i in range(copy_rows[a], copy_rows[b]):
            for j in range(1, 6):
                rangeselectted.append(Summary.cell(row=i, column=j).value)

        num=0
        for x in range(paste_rows[a], paste_rows[b]):
            for y in range(1, 6):
                Test_Summary.cell(row=x, column=y).value = rangeselectted[num]
                num+=1

        rangeselectted.clear()
        a = a + 2
        b = b + 2
        z=z-1
#-------------------------------------------------------------------------------------------------------
    copy_rows = [14, 20, 34, 40, 64, 70, 94, 100, 114, 120]

    a = 0
    b = 1
    items = ["Agency pools", "CDONET", "CHS deals", "HECM pools", "SFW deals"]
    for item in items:
        rangeselectted = []
        for i in range(copy_rows[a], copy_rows[b]):
            for j in range(2, 6):
                rangeselectted.append(Summary.cell(row=i, column=j).value)

        new_sheet = formatted_workbook[item]
        num = 0
        for x in range(7, 13):
            for y in range(18, 22):
                new_sheet.cell(row=x, column=y).value = rangeselectted[num]
                num += 1

        a = a + 2
        b = b + 2
        rangeselectted.clear()

#--------------------------------------------------------------------------------------------------------------
    copy_rows = [44, 50, 74, 80, 124, 130]

    a = 0
    b = 1
    items = ["CDONET", "CHS deals", "SFW deals"]
    for item in items:
        rangeselectted = []
        for i in range(copy_rows[a], copy_rows[b]):
            for j in range(2, 6):
                rangeselectted.append(Summary.cell(row=i, column=j).value)

        new_sheet = formatted_workbook[item]
        num = 0

        for x in range(17, 23):
            for y in range(18, 22):
                new_sheet.cell(row=x, column=y).value = rangeselectted[num]
                num += 1

        a = a + 2
        b = b + 2
        rangeselectted.clear()


    #--------------------------------------------------------------------------------------------------

    raw_sheets = ["Agency_Pools", "CDONET", "CHS", "HECM_Pools", "SFW"]
    formatted_sheets = ["Agency pools", "CDONET", "CHS deals", "HECM pools", "SFW deals"]

    c = 0
    for sheet in raw_sheets:
        sheet2 = raw_workbook[sheet]
        sheet3 = formatted_workbook[formatted_sheets[c]]
        rows = sheet2.max_row
        columns = sheet2.max_column

        if sheet in ["CDONET", "CHS", "SFW"]:
            for row in range(2, rows+1):
                for column in range(2, columns):
                    content = sheet2.cell(row=row, column=column).value
                    sheet3.cell(row=row, column=column-1).value = content
                    currentCell = sheet3.cell(row=row, column=column-1)
                    currentCell.alignment = Alignment(horizontal='center', vertical='center')     #  For center Alignment
        else:
            for row in range(2, rows+1):
                for column in range(1, columns):
                    content = sheet2.cell(row=row, column=column).value
                    sheet3.cell(row=row, column=column).value = content
                    currentCell = sheet3.cell(row=row, column=column)
                    currentCell.alignment = Alignment(horizontal='center', vertical='center')
        c+=1

#----------------------------------------------------------------------------------

    def set_border(ws, cell_range):
        thin = Side(border_style="thin", color="000000")
        for row in ws[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for c in formatted_sheets:
        ws = formatted_workbook[c]
        rows = ws.max_row
        if c in ["Agency pools", "HECM pools"]:
            set_border(ws, f'A1:O{rows}')
        else:
            set_border(ws, f'A1:N{rows}')

    formatted_workbook.save(formatted_file)
    print("Data Copied in ", paste_file_names[k])
    k += 1


#XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX

## Script for Static Model


OS = ["Windows", "Linux"]
copy_file_names = "analysis_results_all_summary_static.xlsx"
paste_file_names = ["MA-SF API_performance_report_static_win64.xlsx", "MA-SF API_performance_report_static_lin64.xlsx"]
k = 0
for oss in OS:
    # Copy
    raw_file = dwndload_res_path+oss+"\\"+copy_file_names
    raw_workbook = openpyxl.load_workbook(raw_file)
    Summary = raw_workbook["Summary"]

    # Paste
    formatted_file = report_path+paste_file_names[k]
    formatted_workbook = openpyxl.load_workbook(formatted_file)
    Test_Summary = formatted_workbook["Test Summary "]


    copy_rows = [4,10,24,30,54,60,84,90,104,110]
    paste_rows = [10,16,20,26,30,36,40,46,50,56]
    a = 0
    b = 1
    z = 5
    while z >= 1:
        rangeselectted = []
        for i in range(copy_rows[a], copy_rows[b]):
            for j in range(1, 6):
                rangeselectted.append(Summary.cell(row=i, column=j).value)

        num=0
        for x in range(paste_rows[a], paste_rows[b]):
            for y in range(1, 6):
                Test_Summary.cell(row=x, column=y).value = rangeselectted[num]
                num+=1

        rangeselectted.clear()
        a = a + 2
        b = b + 2
        z=z-1
#-------------------------------------------------------------------------------------------------------
    copy_rows = [14, 20, 34, 40, 64, 70, 94, 100, 114, 120]

    a = 0
    b = 1
    items = ["Agency pools", "CDONET", "CHS deals", "HECM pools", "SFW deals"]
    for item in items:
        rangeselectted = []
        for i in range(copy_rows[a], copy_rows[b]):
            for j in range(2, 6):
                rangeselectted.append(Summary.cell(row=i, column=j).value)

        new_sheet = formatted_workbook[item]
        num = 0
        for x in range(8, 14):
            for y in range(17, 21):
                new_sheet.cell(row=x, column=y).value = rangeselectted[num]
                num += 1

        a = a + 2
        b = b + 2
        rangeselectted.clear()

#--------------------------------------------------------------------------------------------------------------
    copy_rows = [44, 50, 74, 80, 124, 130]

    a = 0
    b = 1
    items = ["CDONET", "CHS deals", "SFW deals"]
    for item in items:
        rangeselectted = []
        for i in range(copy_rows[a], copy_rows[b]):
            for j in range(2, 6):
                rangeselectted.append(Summary.cell(row=i, column=j).value)

        new_sheet = formatted_workbook[item]
        num = 0

        for x in range(18, 24):
            for y in range(17, 21):
                new_sheet.cell(row=x, column=y).value = rangeselectted[num]
                num += 1

        a = a + 2
        b = b + 2
        rangeselectted.clear()


    #--------------------------------------------------------------------------------------------------

    raw_sheets = ["Agency_Pools", "CDONET", "CHS", "HECM_Pools", "SFW"]
    formatted_sheets = ["Agency pools", "CDONET", "CHS deals", "HECM pools", "SFW deals"]

    c = 0
    for sheet in raw_sheets:
        sheet2 = raw_workbook[sheet]
        sheet3 = formatted_workbook[formatted_sheets[c]]
        rows = sheet2.max_row
        columns = sheet2.max_column

        if sheet in ["CDONET", "CHS", "SFW"]:
            for row in range(2, rows+1):
                for column in range(2, columns):
                    content = sheet2.cell(row=row, column=column).value
                    sheet3.cell(row=row, column=column-1).value = content
                    currentCell = sheet3.cell(row=row, column=column-1)
                    currentCell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            for row in range(2, rows+1):
                for column in range(1, columns):
                    content = sheet2.cell(row=row, column=column).value
                    sheet3.cell(row=row, column=column).value = content
                    currentCell = sheet3.cell(row=row, column=column)
                    currentCell.alignment = Alignment(horizontal='center', vertical='center')
        c+=1

#------------------------------------------------------------------------

    for c in formatted_sheets:
        ws = formatted_workbook[c]
        rows = ws.max_row
        if c in ["Agency pools", "HECM pools"]:
            set_border(ws, f'A1:N{rows}')
        else:
            set_border(ws, f'A1:M{rows}')

    formatted_workbook.save(formatted_file)
    print("Data Copied in ", paste_file_names[k])
    k += 1




